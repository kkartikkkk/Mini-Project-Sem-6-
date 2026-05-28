import pickle, re, os

import openpyxl

from sklearn.model_selection import train_test_split

from sklearn.metrics import classification_report, accuracy_score

BASE_DIR = os.path.dirname(__file__)


def load(name):
    with open(os.path.join(BASE_DIR, name), "rb") as f:
        return pickle.load(f)

cat_model  = load("category_model.pkl")

cat_le     = load("category_encoder.pkl")

prio_model = load("priority_model.pkl")

prio_le    = load("priority_encoder.pkl")

candidates = [

    os.path.join(BASE_DIR, "..", "ecommerce_consumer_complaints_5000.xlsx"),

    os.path.join(BASE_DIR, "ecommerce_consumer_complaints_5000.xlsx"),

]

path = next((p for p in candidates if os.path.exists(p)), None)

if not path:
    raise FileNotFoundError("Dataset not found.")

wb      = openpyxl.load_workbook(path)

ws      = wb.active

headers = [c.value for c in ws[1]]

cat_idx  = headers.index("Category")

text_idx = headers.index("Complaint_Text")

prio_idx = headers.index("Priority_Label")

texts, cats, prios = [], [], []

for row in ws.iter_rows(min_row=2, values_only=True):
    t, c, p = row[text_idx], row[cat_idx], row[prio_idx]

    if t and c and p:
        texts.append(re.sub(r"[^a-z\s]", "", str(t).lower()))

        cats.append(c)

        prios.append(p)

print(f"Dataset: {len(texts)} samples\n")

y_cat = cat_le.transform(cats)

X_tr, X_te, y_tr, y_te = train_test_split(
    texts, y_cat, test_size=0.2, stratify=y_cat, random_state=42

)

train_acc = accuracy_score(y_tr, cat_model.predict(X_tr))

test_acc  = accuracy_score(y_te, cat_model.predict(X_te))

print("=" * 55)

print("  CATEGORY MODEL  (TF-IDF + Logistic Regression)")

print("=" * 55)

print(f"  Training Accuracy : {train_acc*100:.2f}%")

print(f"  Testing  Accuracy : {test_acc*100:.2f}%")

print("\n  Per-class report (test set):")

print(classification_report(y_te, cat_model.predict(X_te),

                             target_names=cat_le.classes_))

y_prio = prio_le.transform(prios)

X_tr2, X_te2, yp_tr, yp_te = train_test_split(
    texts, y_prio, test_size=0.2, stratify=y_prio, random_state=42

)

train_acc2 = accuracy_score(yp_tr, prio_model.predict(X_tr2))

test_acc2  = accuracy_score(yp_te, prio_model.predict(X_te2))

print("=" * 55)

print("  PRIORITY MODEL   (TF-IDF + Logistic Regression)")

print("=" * 55)

print(f"  Training Accuracy : {train_acc2*100:.2f}%")

print(f"  Testing  Accuracy : {test_acc2*100:.2f}%")

print("\n  Per-class report (test set):")

print(classification_report(yp_te, prio_model.predict(X_te2),

                             target_names=prio_le.classes_))
