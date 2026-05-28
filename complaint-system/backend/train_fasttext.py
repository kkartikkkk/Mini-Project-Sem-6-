import os

import re

import random

BASE_DIR   = os.path.dirname(__file__)

PARENT_DIR = os.path.dirname(BASE_DIR)

DATASET_CANDIDATES = [

    os.path.join(PARENT_DIR, "ecommerce_consumer_complaints_5000.xlsx"),

    os.path.join(BASE_DIR,   "ecommerce_consumer_complaints_5000.xlsx"),

]

DATASET_PATH = next((p for p in DATASET_CANDIDATES if os.path.exists(p)), None)

if not DATASET_PATH:
    raise FileNotFoundError("Dataset file 'ecommerce_consumer_complaints_5000.xlsx' not found.")

print(f"Dataset: {DATASET_PATH}")

LABEL_MAP = {
    "Product Issue":    "Product_Issue",
    "Wrong Item":       "Product_Issue",
    "Customer Service": "Customer_Service",
    "Delivery Issue":   "Delivery_Issue",
    "Refund Request":   "Refund_Request",
    "Account Issue":    "Account_Issue",
    "Product Inquiry":  "Product_Inquiry",
    "Payment Issue":    "Payment_Issue",
}

import openpyxl

wb = openpyxl.load_workbook(DATASET_PATH)

ws = wb.active

headers = [c.value for c in ws[1]]

cat_idx  = headers.index("Category")

text_idx = headers.index("Complaint_Text")

rows = []

for row in ws.iter_rows(min_row=2, values_only=True):
    cat  = row[cat_idx]

    text = row[text_idx]

    if not cat or not text:
        continue

    label = LABEL_MAP.get(cat, cat.replace(" ", "_"))

    clean = re.sub(r"\s+", " ", str(text).lower().strip())

    rows.append((label, clean))

print(f"Loaded {len(rows)} samples")

random.seed(42)

random.shuffle(rows)

split = int(len(rows) * 0.9)

train_rows = rows[:split]

valid_rows = rows[split:]

TRAIN_FILE = os.path.join(BASE_DIR, "ft_train.txt")

VALID_FILE = os.path.join(BASE_DIR, "ft_valid.txt")

MODEL_PATH = os.path.join(BASE_DIR, "complaint_fasttext.bin")


def write_ft_file(path, data):
    with open(path, "w", encoding="utf-8") as f:
        for label, text in data:
            f.write(f"__label__{label} {text}\n")

write_ft_file(TRAIN_FILE, train_rows)

write_ft_file(VALID_FILE, valid_rows)

print(f"Train: {len(train_rows)}  |  Valid: {len(valid_rows)}")

try:
    import fasttext

except ImportError:
    print("\nFastText not installed. Run:  pip install fasttext-wheel")

    raise

print("\nTraining FastText model...")

model = fasttext.train_supervised(
    input      = TRAIN_FILE,

    epoch      = 50,

    lr         = 0.5,

    wordNgrams = 2,

    dim        = 100,

    loss       = "softmax",

    verbose    = 2,

)

result = model.test(VALID_FILE)

n, precision, recall = result

print(f"\nValidation:  N={n}  P@1={precision:.4f}  R@1={recall:.4f}")

model.save_model(MODEL_PATH)

print(f"\nModel saved → {MODEL_PATH}")

print("Restart the backend to activate FastText classification.")

os.remove(TRAIN_FILE)

os.remove(VALID_FILE)
